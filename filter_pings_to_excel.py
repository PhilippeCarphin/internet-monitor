#!/usr/bin/env python3

import os
import sys
import datetime
import time
import re
import argparse
# import pyxlsx
import openpyxl
import pandas as pd

def get_arguments():

    p = argparse.ArgumentParser(description="Monitor my god damn internet connection because it seems that stupid Videotron is unable to do it themselves")

    p.add_argument("--show-uptimes", action='store_true')
    p.add_argument("--alerts", action='store_true')

    return p.parse_args()

def get_time(log_line):
    """
    Lines are of the form
        Internet working at 2022-10-27 22:08:33
    or
        Internet down at 2022-10-27 22:08:33
    so words 3 and 4 make up the date.
    """
    words = log_line.split()
    time_string = f"{words[3]} {words[4]}"
    dt = datetime.datetime.strptime(time_string, "%Y-%m-%d %H:%M:%S")
    return dt


def color_out_duration(d):
    if d < datetime.timedelta(seconds=30):
        color = bad_colors[5]
    elif d < datetime.timedelta(minutes=2):
        color = bad_colors[4]
    elif d < datetime.timedelta(minutes=5):
        color = bad_colors[3]
    elif d < datetime.timedelta(minutes=20):
        color = bad_colors[2]
    elif d < datetime.timedelta(minutes=40):
        color = bad_colors[1]
    else:
        color = bad_colors[0]
    return f"{color}{d}\033[0m"

def color_up_duration(d):
    if d < datetime.timedelta(seconds=30):
        color = bad_colors[4]
    elif d < datetime.timedelta(minutes=1):
        color = good_colors[0]
    elif d < datetime.timedelta(minutes=5):
        color = good_colors[1]
    elif d < datetime.timedelta(minutes=10):
        color = good_colors[2]
    elif d < datetime.timedelta(minutes=40):
        color = good_colors[3]
    elif d < datetime.timedelta(minutes=120):
        color = good_colors[4]
    else:
        color = good_colors[5]
    return f"{color}{d}\033[0m"

bad_colors = [
        "\033[1;38;5;196m", # red
        "\033[1;38;5;202m",
        "\033[1;38;5;208m",
        "\033[1;38;5;214m",
        "\033[1;38;5;220m",
        "\033[1;38;5;226m", # yellow
        ]
good_colors = [
        "\033[1;38;5;226m", # yellow
        "\033[1;38;5;190m",
        "\033[1;38;5;154m",
        "\033[1;38;5;118m",
        "\033[1;38;5;82m",
        "\033[1;38;5;46m", # Green
]

def stdin_to_dataframe():

    args = get_arguments()

    state = "working"
    start = sys.stdin.readline()
    interval = {'start': get_time(start)}
    rows = []
    cutoff = 1000
    i = 0
    for l in sys.stdin:
        # So I can temporarily remove some lines in the test log
        if l.strip().startswith('#'):
            continue
        ltime = get_time(l)
        #
        # Working to working
        #
        if state == "working" and "working" in l:
            pass
        #
        # Down to down
        #
        elif state == "down" and "down" in l:
            pass
        #
        # Down to working
        #
        elif state == "down" and "working" in l:
            new_row = {"Status": "OUT", "Date": interval['start'], "Duration": ltime - interval['start']}
            rows.append(new_row)
            interval = {'start': get_time(l)}
            state = "working"
        #
        # Working to down
        #
        elif state == "working" and "down" in l:
            new_row = {"Status": "UP", "Date": interval['start'], "Duration": ltime - interval['start']}
            rows.append(new_row)
            state = "down"
            interval = {'start': get_time(l)}

        i += 1
        # if i > cutoff:
        #     break

    return pd.DataFrame(rows)

def df_to_excel(df):
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.append(["Status", "Date", "Duration"])
    row_number = 1 # Row numbers start at 1
    for row in df.iterrows():
        row_number += 1
        series = row[1]
        status, date, duration = series['Status'], series['Date'], series['Duration']
        xl_row = [status, date.strftime("%y-%m-%d %H:%M:%S"), f"{duration}"]
        ws.append(xl_row)
        status_cell = ws[f'A{row_number}']
        duration_cell = ws[f'C{row_number}']
        color = None
        if status == "UP":
        # - Append to worksheet
            status_cell.font = openpyxl.styles.Font(color='FF00FF00')
            color = up_duration_to_hex_color(duration)
        else:
            status_cell.font = openpyxl.styles.Font(color='FFFF0000')
            color = out_duration_to_hex_color(duration)
        print(f"{status:3}, {date}, {duration}, {color}")
        fill = openpyxl.styles.PatternFill(fill_type='solid', start_color=color, end_color=color)
        duration_cell.fill = fill
        # - Color last row based on duration
    return wb

heatmap_colors = [
    'FF488F31', #0
    'FF6B9C38', #1
    'FF8BA843', #2
    'FFAAB450', #3
    'FFC7C05F', #4
    'FFE3CC71', #5
    'FFFFD885', #6
    'FFFDC173', #7
    'FFFBA965', #8
    'FFF7915C', #9
    'FFF17858', #10
    'FFE95E58', #11
    'FFDE425B', #12
]

def up_duration_to_hex_color(d):
    if d < datetime.timedelta(seconds=30):
        color = heatmap_colors[5]
    elif d < datetime.timedelta(minutes=2):
        color = heatmap_colors[4]
    elif d < datetime.timedelta(minutes=5):
        color = heatmap_colors[3]
    elif d < datetime.timedelta(minutes=20):
        color = heatmap_colors[2]
    elif d < datetime.timedelta(minutes=40):
        color = heatmap_colors[1]
    else:
        color = heatmap_colors[0]
    return color

def out_duration_to_hex_color(d):
    if d < datetime.timedelta(seconds=30):
        color = heatmap_colors[4]
    elif d < datetime.timedelta(minutes=1):
        color = heatmap_colors[5]
    elif d < datetime.timedelta(minutes=5):
        color = heatmap_colors[8]
    elif d < datetime.timedelta(minutes=10):
        color = heatmap_colors[9]
    elif d < datetime.timedelta(minutes=40):
        color = heatmap_colors[10]
    elif d < datetime.timedelta(minutes=120):
        color = heatmap_colors[11]
    else:
        color = heatmap_colors[12]
    return color

def main():

    df = stdin_to_dataframe()

    xl = df_to_excel(df)

    print(f" worksheets: {xl.worksheets}")
    xl.save("internet_outages.xlsx")

    # xlsx = df.to_excel("tmp_excel_file.xlsx")

    print(df.head())

try:
    main()
except KeyboardInterrupt:
    pass
